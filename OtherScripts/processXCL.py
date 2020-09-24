# from openpyxl.styles import colors
# from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from json import dumps

wbname = 'LCCWCatTest.xlsx'
wb = load_workbook(filename = wbname)
sh = wb.active
colors = []
cobjs = []
ccells = []
for rw in sh.iter_rows(1, 750):
    cl = rw[0]
    ccells.append(cl)
    rgb = cl.fill.fgColor.rgb
    if rgb not in colors:
        colors.append(rgb)
        cobj = {
            "color": rgb,
            "rnum": cl.row,
            "rtitle": rw[1].value
        }
        cobjs.append(cobj)

print(dumps(cobjs, ensure_ascii=False))
print(ccells)
# cl = sh.cell(5, 1)
# print(cl.fill.fgColor.rgb)
# cl = sh.cell(1, 1)
# print(cl.fill.fgColor)
# cl = sh.cell(10, 2)
# print(cl.fill.fgColor)
